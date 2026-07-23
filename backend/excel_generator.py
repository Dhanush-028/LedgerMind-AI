"""
excel_generator.py

Builds an .xlsx export of the cash-flow forecast.

Install: pip install openpyxl
"""
import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BUCKET_LABELS = {"0-30": "Next 30 days", "31-60": "31-60 days", "61-90": "61-90 days", "90+": "90+ days"}

HEADER_FILL = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True)
BOLD = Font(bold=True)


def generate_forecast_excel(projection: dict, narrative: dict, as_of: date) -> io.BytesIO:
    wb = Workbook()

    # --- Sheet 1: Forecast table -------------------------------------------------
    ws = wb.active
    ws.title = "Forecast"

    ws["A1"] = "LedgerMind AI — Cash-Flow Forecast"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated {as_of.isoformat()}"
    ws["A2"].font = Font(color="6b7280")

    total = sum(projection.values())
    ws["A4"] = "Total expected inflow (risk-weighted)"
    ws["A4"].font = BOLD
    ws["B4"] = total
    ws["B4"].number_format = '"₹"#,##0'

    headers = ["Window", "Expected Inflow (INR)"]
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for key in ["0-30", "31-60", "61-90", "90+"]:
        ws.append([BUCKET_LABELS.get(key, key), projection.get(key, 0)])
        ws.cell(row=ws.max_row, column=2).number_format = '"₹"#,##0'

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    # --- Sheet 2: AI Narrative -----------------------------------------------------
    ws2 = wb.create_sheet("AI Narrative")
    ws2["A1"] = "Summary"
    ws2["A1"].font = BOLD
    ws2["A2"] = narrative.get("summary", "")
    ws2["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[2].height = 60

    row = 4
    risks = narrative.get("risks") or []
    if risks:
        ws2.cell(row=row, column=1, value="Risks").font = BOLD
        row += 1
        for r in risks:
            ws2.cell(row=row, column=1, value=f"• {r}")
            ws2.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            row += 1
        row += 1

    recs = narrative.get("recommendations") or []
    if recs:
        ws2.cell(row=row, column=1, value="Recommendations").font = BOLD
        row += 1
        for r in recs:
            ws2.cell(row=row, column=1, value=f"• {r}")
            ws2.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            row += 1

    ws2.column_dimensions["A"].width = 90

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer